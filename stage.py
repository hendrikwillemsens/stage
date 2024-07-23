import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import openpyxl
import math
import statistics
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import time

start_time = time.time()
def print_elapsed_time(start_time, message):
    elapsed_time = time.time() - start_time
    print(f"{message} - Tijd verstreken sinds start: {elapsed_time:.2f} seconden")

# Laad het geüploade bestand en het gedownloade bestand
geüpload_bestand = st.file_uploader("Kies een bestand (xlsx) om te tellen", type='xlsx')
download_bestand = st.file_uploader("Kies een bestand (xlsx) om het resultaat te schrijven", type='xlsx')
analyse_bestand = st.file_uploader("Kies een bestand (xlsx) om te analyseren", type='xlsx')
file_name = st.text_input('Voer de gewenste bestandsnaam in (zonder extensie):', 'statistics')
zoekwaarde = st.text_input("Voer een waarde in om naar te zoeken (bijv. 473-1 of alles)", key="zoekwaarde")

if geüpload_bestand and download_bestand and analyse_bestand and file_name and zoekwaarde:
             
        wb_lezen = load_workbook(filename=BytesIO(geüpload_bestand.read()), data_only=True)
        time.sleep(2)
        print_elapsed_time(start_time, "Na load workbook")
        wb_schrijven = load_workbook(filename=BytesIO(download_bestand.read()))
        time.sleep(2)
        print_elapsed_time(start_time, "Na load workbook")
        wb_analyseren = load_workbook(filename=BytesIO(analyse_bestand.read()))
        time.sleep(2)
        print_elapsed_time(start_time, "Na load workbook")
        
        
        

        kolom = "G"
        ws_lezen = wb_lezen.active

        def hoofden(werbladx):
            werbladx['A1'] = "Onze ref."
            werbladx['B1'] = "Naam"
            werbladx['C1'] = "Voornaam"
            werbladx['D1'] = "Bedrijf"
            werbladx['E1'] = "Werknemer"
            werbladx['F1'] = "Ontvangstdatum"
            werbladx['G1'] = "Geboortedatum"
            werbladx['H1'] = "Geslacht"
            werbladx['I1'] = "Arts"
            bold_font = Font(bold=True)    
            for cell in werbladx[1]:
                cell.font = bold_font

        def custom_sort_key(entry):
            parts = entry.split()
            if len(parts) > 1:
                base_name = parts[0]
                is_creat = 'creat.' in entry.lower()
            else:
                base_name = entry
                is_creat = False
            return (base_name, not is_creat) 
        
        detectielimieten_analysen={"1-methoxy-2-propanol": 0.5 ,"Chroom":0.5, "Cobalt":0, "Ethylmethylketon":0.1,"Fluor":0, "Hippuurzuur":20, "Methylhippuurzuur":50.0,
        "Muconzuur":0, "Nikkel":0, "o-Cresol":0.05, "Creatinine":0}
        
        nieuwe_limieten={"1-methoxy-2-propanol": 0.4 ,"Chroom":0.4, "Cobalt":0, "Ethylmethylketon":0.1,"Fluor":0, "Hippuurzuur":19, "Methylhippuurzuur":49.9,
        "Muconzuur":0, "Nikkel":0, "o-Cresol":0.04, "Creatinine":0}
        
        tollerantiegrens = {"1-methoxy-2-propanol":"niet beschikbaar","Chroom": {"Einde shift, einde werkweek":"30,0 µg/g creat.",
        "Einde shift - begin shift": "</=10,0 µg/g creat."},
        "Cobalt":{"Einde shift, einde werkweek":"15,0 µg/g creat." }, "Ethylmethylketon":{"Einde shift":"2,5 mg/g creat."},
        "Fluor":{"Begin shift":"3,00 mg/g creat.", "Einde shift": "7,00 mg/g creat."}, "Hippuurzuur":{" Einde shift": "1500 mg/g creat."}, 
        "Methylhippuurzuur": {"Einde shift": "1500,0 mg/g creat."}, "Muconzuur":{"Einde shift":"1,50 mg/g creat."},
        "Nikkel": ["Blootstelling aan oplosbaar nikkel (100 µg/m3 lucht) correspondeert einde shift met", "50,0 µg/g creat."],
        "o-Cresol":{"Einde shift": "0,50 mg/g creat."}                                                                                                                                    
        }
        
        grenswaarden = {"1-methoxy-2-propanol":"0","Chroom": "0,0 - 0,35 µg/g creat.",
        "Cobalt":"0,0 - 2,0 µg/g creat.", "Ethylmethylketon":"0",
        "Fluor":"0,00 - 1,00 mg/g creat.", "Hippuurzuur": "0 - 1500 mg/g creat.", 
        "Methylhippuurzuur": "0", "Muconzuur":" 0,00 - 0,30 mg/g creat.",
        "Nikkel": "0,0 - 5,0 µg/g creat.",
        "o-Cresol":"0,00 - 0,30 mg/g creat."                                                                                                                                    
        }

        rijen_kopie = []
        for rij in ws_lezen.iter_rows(min_row=2, values_only=True):
            if rij[6] == zoekwaarde and rij[15] != 0:
                rijen_kopie.append(rij)
            elif zoekwaarde == "alles":
                rijen_kopie.append(rij)

        ws_schrijven = wb_schrijven.active

        kolommen =  [0,1,2,6,8,9,10,11,20,12,14,15]
        kolommen2 = [0,1,2,3,4,5,6,7,8,9,11,12]
        
        ws_schrijven = wb_schrijven.active
        for index, rij in enumerate(rijen_kopie, start=2):
            for i, kolom_index in enumerate(kolommen):
                waarde = rij[kolom_index]
                doelkolom_index = kolommen2[i] 
                ws_schrijven.cell(row=index, column=doelkolom_index + 1, value=waarde)
        
        analyseid = []
        for rij in ws_schrijven.iter_rows(min_row = 2, values_only=True):
            if rij[9] is not None:
                analyseid.append(rij[9])
        time.sleep(2)
        print_elapsed_time(start_time, "voor wb_analyseren")
        
        
        ws_analyse = wb_analyseren.active
        hoeveelheid = []
        alle_urine_analysen = []
        alle_urine_analysen2 = []
        for row in analyseid:
            for rij in ws_analyse.iter_rows(min_row=2, values_only=True):
                if rij[0] == row:
                    value_B = rij[1]  
                    value_C = rij[2]  
                    concatenated_value = f"{value_B} ({value_C})"  
                    hoeveelheid.append(concatenated_value)
                    if rij[6] ==  "Urine" and concatenated_value not in alle_urine_analysen:
                        alle_urine_analysen.append(concatenated_value)
                        alle_urine_analysen.append("")
                        
                    if rij[6] ==  "Urine" and value_B not in alle_urine_analysen2:
                        alle_urine_analysen2.append(value_B)
                            
                    break  

        
        # Tweede timestamp
        time.sleep(2)
        print_elapsed_time(start_time, "Na het verwerken van alle analyseid")

        ws_schrijven = wb_schrijven.active
        kolom_letter = 'K' 
        begin_rij = 2 

        for index, waarde in enumerate(hoeveelheid):
            cel = ws_schrijven[kolom_letter + str(begin_rij + index)] 
            cel.value = waarde
        analyses = []

        
        
        header = [cell.value for cell in ws_schrijven[1]]
        patient_id_index = header.index("Onze ref.") 
        analyse_index = header.index("Analyse")
        T_index = header.index("T")
        result_index = header.index("Resultaat")
        other_columns = [i for i in range(len(header)) if i not in [patient_id_index, analyse_index, result_index]]

        

        patients = dict()
        t = dict()
        analyses_unique = []
        analyses_unique2 = []

        for row in ws_schrijven.iter_rows(min_row=2, values_only=True):
            patient_id = str(row[patient_id_index]) 
            analyse = row[analyse_index]
            result = row[result_index]
            if analyse and analyse not in analyses_unique:
                analyses_unique.append(analyse)

            if analyse:
               delen = analyse.split('(')
               eerste_deel = delen[0].strip()
        

               if eerste_deel and eerste_deel not in analyses_unique2:
                   analyses_unique2.append(eerste_deel)

        
            if patient_id not in patients:
                patients[patient_id] = dict()

                for i in other_columns:
                    patients[patient_id][header[i]] = row[i]
            patients[patient_id][analyse] = result
            
            T_cell = row[T_index]
            if T_cell is not None:
            
                if patient_id not in t:
                    t[patient_id] = dict()
                t[patient_id][analyse] = T_cell
                
            
       
        time.sleep(2)
        print_elapsed_time(start_time, "Na dictionary")
        non_blank_entries = sorted([entry for entry in analyses_unique if entry != "" and entry in alle_urine_analysen], key=custom_sort_key)
        sorted_urine_analysen = []
        for entry in non_blank_entries:
            sorted_urine_analysen.append(entry)
            sorted_urine_analysen.append("")


        # Replace the original list with the sorted list
        urine_analysen = sorted_urine_analysen
        urine_analysen22 = [entry for entry in analyses_unique2 if entry != "" and entry in alle_urine_analysen2]
        urine_analysen2 = urine_analysen22

        # Filter beschikbare analyses
        filter_analyses = [analyse for analyse in urine_analysen2 if analyse not in detectielimieten_analysen]

        # Session state initialisatie
        if 'analyses' not in st.session_state:
            st.session_state.analyses = filter_analyses.copy()
        if 'toegevoegde_analyses' not in st.session_state:
            st.session_state.toegevoegde_analyses = {}

        # Multiselect voor geselecteerde analyses om te negeren
        geselecteerde_analyses = st.multiselect("Selecteer analyses om niet in te vullen:", st.session_state.analyses)

        # Dictionary voor ingevulde waarden
        ingevulde_hoeveelheden = {}

        # Invoer voor ontbrekende analyses
        for analyse in st.session_state.analyses:
            if analyse not in geselecteerde_analyses:
                hoeveelheid = st.text_input(f"Voer de hoeveelheid in voor {analyse}:", key=analyse)
                if hoeveelheid:
                    ingevulde_hoeveelheden[analyse] = hoeveelheid

        # Knop om analyses toe te voegen
        if st.button("Voeg analyses toe"):
            for analyse, hoeveelheid in ingevulde_hoeveelheden.items():
                detectielimieten_analysen[analyse] = detectielimieten_analysen.get(analyse, 0) + int(hoeveelheid)
                st.session_state.toegevoegde_analyses[analyse] = hoeveelheid

            # Verwijder geselecteerde analyses uit de analyses en beschikbare analyses
            st.session_state.analyses = [analyse for analyse in st.session_state.analyses if analyse not in geselecteerde_analyses]
            urine_analysen2 = [analyse for analyse in urine_analysen2 if analyse not in geselecteerde_analyses]
    


        time.sleep(3)
        print_elapsed_time(start_time, "Na derde taak")
        filtered_analysen = []
        for analyse in urine_analysen:
            if analyse == "":
                continue
            analyse_name = analyse.split(" ")[0]  # Neem alleen het deel voor de eerste spatie
            if analyse_name in urine_analysen2:
                filtered_analysen.append(analyse)



        sorted_filtered_analysen = []
        for entry in filtered_analysen:
            sorted_filtered_analysen.append(entry)
            sorted_filtered_analysen.append("")

        urine_analysen = sorted_filtered_analysen




        time.sleep(3)
        print_elapsed_time(start_time, "voor resultaten")
        # write data to sheet, create new sheet if not exists, empty sheet if exists
        if "Resultaten" not in wb_schrijven.sheetnames:
            ws2 = wb_schrijven.create_sheet(title="Resultaten")
            # write header
            ws2.append(["Onze ref."]+(header:=["Naam", "Voornaam", "Bedrijf", "Werknemer", "Ontvangstdatum", "Geboortedatum", "Geslacht", "Arts", " "])+urine_analysen)
        else:
            ws2 = wb_schrijven["Resultaten"]
            ws2.delete_rows(2, ws2.max_row)

        for patient_id, patient_data in patients.items():
   
                row = [patient_id]+[patient_data.get(header[i], "") for i in range(len(header))]
                analyse_array = []
                
                for analyse in urine_analysen:
                    analyse_array.append(patient_data.get(analyse, ""))
                if any(element != "" for element in analyse_array):
                    rij = row + analyse_array
                    ws2.append(rij)

        rijen_index = 2
        
        for row_index, row in enumerate(ws2.iter_rows(min_row=2), start=2):
            first_cell_value = row[0].value
            
            for id_keys, id_value in t.items():
                if first_cell_value == id_keys:  
                    for cell_index, cell in enumerate(row, start=1):  
                        kolommen_index = get_column_letter(cell_index)
                        
                        waarde = ws2[f'{kolommen_index}1'].value
                        if waarde in id_value:
                            echte_kolom = cell_index - 1
                            echte_kolommen_index = get_column_letter(echte_kolom)
                            
                            ws2[f'{echte_kolommen_index}{rijen_index}'] = "<"
            rijen_index += 1

        for column in ws2.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:  # Als de celwaarde geen string is
                    pass
            adjusted_width = (max_length + 2)
            ws2.column_dimensions[column[0].column_letter].width = adjusted_width           

        

            
        for analyse in urine_analysen2:
            if analyse not in wb_schrijven.sheetnames and analyse != "Creatinine":
                ws2 = wb_schrijven.create_sheet(title=analyse)
                detectielimiet = detectielimieten_analysen.get(analyse,"")
                nieuwe_liemiet = nieuwe_limieten.get(analyse,"")
                
                headers = ["Onze ref.", "Naam", "Voornaam", "Bedrijf", "Werknemer", "Ontvangstdatum", "Geboortedatum", "Geslacht", "Arts", " "]
                hoofd = ["Naam", "Voornaam", "Bedrijf", "Werknemer", "Ontvangstdatum", "Geboortedatum", "Geslacht", "Arts", " "]
                vermijd_analyses = ["methylhippuurzuur (mg/L)", "methylhippuurzuur (mg/g creat.)"]

                if analyse != "Methylhippuurzuur": 
                    matching_headers = [result for result in urine_analysen if analyse in result and all(v_analyse not in result for v_analyse in vermijd_analyses)]
                elif analyse == "Methylhippuurzuur":
                    matching_headers = [result for result in urine_analysen if analyse in result]
                match = matching_headers
                
                
            
                formatted_headers = []
                for header in matching_headers:
                    formatted_headers.append(header)
                    formatted_headers.append("")  # Voeg een lege kolom toe na elk item
                headers += formatted_headers
                ws2.append(headers)  # Voeg de headers toe aan het werkblad
                
                
                 
                
                last_filled_cell = None
                for cell in ws2[1]:  # ws[1] gets the first row
                    if cell.value is not None:
                        last_filled_cell = cell
                
                
                if last_filled_cell:
                    next_col = last_filled_cell.column + 1
                else:
                    next_col = 1

                ws2.cell(row=1, column=next_col, value="Creatinine (g/L)")
                
               
                c = 1
                for patient_id, patient_data in patients.items():
                    row_data = []
                    all_data_present = True  # Gebruik een vlag om te controleren of alle data aanwezig is voor de huidige patiënt
 
                    row = [patient_id] + [patient_data.get(col, "") for col in hoofd]
                    for heading in matching_headers:
                        if heading in patient_data:
                            substance_value = patient_data[heading]
                            row = row + [substance_value] + [""]
                            
                    c = c + 1
                    ws2.append(row)
                
                
                
                rij1= [cell.value for cell in ws2[1]]
                creatinine_index = rij1.index("Creatinine (g/L)")
                creatinine_index = creatinine_index + 1
                for i in ws2['K'][1:]:
                    if i.value == None:
                        ws2.delete_rows(i.row)
                        c = c - 1
                
                rij = 2
                for row in ws2.iter_rows(min_row=2, values_only=True):  
                    cel_waarde = ws2[f'A{rij}'].value
                    
                    creatinine_value = patients[cel_waarde]['Creatinine (g/L)']
                    
                    ws2.cell(row=rij, column=creatinine_index, value=creatinine_value)
                    rij = rij + 1

                       
                gemiddeld = c + 1
                aantal_stalen = c - 1     
                som = 0
                detectie = str(detectielimiet)
                if '.' in detectie:
                    decimal_part = detectie.split('.')[1]
                    num_decimals = len(decimal_part)
                    minima = detectielimiet - (10 ** -num_decimals)
                else:
                    decimals = 0
                    minima = detectielimiet - 1
                    if detectielimiet == 0:
                        minima = 0

                lijst = []
                
                for index, row in enumerate(ws2.iter_rows(values_only = True), start = 1):
                    
                    if index > 1 and index < gemiddeld:
                        
                        if row[12] > detectielimiet:
                            som = som + row[10]
                            lijst.append(row[10])
                        else:
                            som = som + minima
                            lijst.append(minima)
                            ws2[f'J{index}'].value = "<"
                            ws2[f'L{index}'].value = "<"
                
                plt.figure(figsize=(10, 6))
                plt.boxplot(lijst)
                plt.title('Boxplot van Gegevens')
                plt.ylabel('Waarde')
                
                box = "boxplot_" + analyse + ".png"
                plt.savefig(box)
                plt.close()
                img = Image(box)
                
                ws2.add_image(img, f'K{c + 20}')

                aantal_stalen = c - 1    
                average = som/aantal_stalen
                    
                

                statistieken = {}

                # Bereken standaarddeviatie en standaardfout
                try:
                    std_dev_sample = statistics.stdev(lijst)
                    statistieken['standaarddeviatie'] = std_dev_sample
                    standard_error = std_dev_sample / math.sqrt(aantal_stalen)
                    statistieken['standaardfout'] = standard_error
                except statistics.StatisticsError as e:
                    print(f"Fout bij het berekenen van standaarddeviatie of standaardfout: {e}")
                    std_dev_sample = "niet beschikbaar"
                    standard_error = "niet beschikbaar"

                # Bereken mediaan
                try:
                    mediaan = statistics.median(lijst)
                    statistieken['mediaan'] = mediaan
                except statistics.StatisticsError as e:
                    print(f"Fout bij het berekenen van mediaan: {e}")
                    mediaan = "niet beschikbaar"

                # Bereken modus
                try:
                    modus = statistics.mode(lijst)
                    statistieken['modus'] = modus
                except statistics.StatisticsError as e:
                    modus = "niet beschikbaar"

                # Bereken steekproefvariantie
                try:
                    sample_variance = statistics.variance(lijst)
                    statistieken['steekproefvariantie'] = sample_variance
                except statistics.StatisticsError as e:
                    sample_variance = "niet beschikbaar"

                # Bereken bereik
                try:
                    data_range = max(lijst) - min(lijst)
                    statistieken['bereik'] = data_range
                except ValueError as e:
                    data_range= "niet beschikbaar"
                # Bereken minimum
                try:
                    minimum = min(lijst)
                    statistieken['minimum'] = minimum
                except ValueError as e:
                    minimum= "niet beschikbaar"

                # Bereken maximum
                try:
                    maximum = max(lijst)
                    statistieken['maximum'] = maximum
                except ValueError as e:
                    maximum= "niet beschikbaar"

                # Bereken som
                try:
                    som = sum(lijst)
                    statistieken['som'] = som
                except TypeError as e:
                    som= "niet beschikbaar"

                if std_dev_sample != 0 and aantal_stalen not in (1,2,3):
                    deel_1 = (aantal_stalen * (aantal_stalen + 1)) / ((aantal_stalen - 1) * (aantal_stalen - 2) * (aantal_stalen - 3))   #kurtosis
                    deel_2 = sum(((x - average) / std_dev_sample) ** 4 for x in lijst)
                    deel_1_2 = deel_1 * deel_2
                    deel_3 = (3 * (aantal_stalen - 1) ** 2) / ((aantal_stalen - 2) * (aantal_stalen - 3))
                    data_kurtosis = deel_1_2 - deel_3
                else:
                    data_kurtosis = "kurtosis is niet beschikbaar"                                                 
                

                if data_kurtosis != "kurtosis is niet beschikbaar":                                                  #scheefheid
                    deel_4 = sum(((x - average) / std_dev_sample) ** 3 for x in lijst)
                    deel5 = (aantal_stalen) / ((aantal_stalen - 1) * (aantal_stalen - 2)) 
                    data_skewness = deel_4 * deel5 
                else:
                    data_skewness = "scheefheid is niet beschikbaar"


                value_M1 = ws2['M1'].value
                value_K1 = ws2['K1'].value
                
                ws2[f'B{c + 10}'].value = average
                ws2[f'B{c + 11}'].value = std_dev_sample
                ws2[f'B{c + 12}'].value = standard_error
                ws2[f'B{c + 13}'].value = mediaan
                ws2[f'B{c + 14}'].value = modus
                ws2[f'B{c + 15}'].value = sample_variance
                ws2[f'B{c + 16}'].value = data_kurtosis
                ws2[f'B{c + 17}'].value = data_skewness
                ws2[f'B{c + 18}'].value = data_range
                ws2[f'B{c + 19}'].value = minimum
                ws2[f'B{c + 20}'].value = maximum
                ws2[f'B{c + 21}'].value = som
                ws2[f'B{c + 22}'].value = aantal_stalen


                ws2[f'A{c + 8}'].value = value_K1
                ws2[f'A{c + 10}'].value = "average"
                ws2[f'A{c + 11}'].value = "standaarddeviatie"
                ws2[f'A{c + 12}'].value = "standaardfout"
                ws2[f'A{c + 13}'].value = "mediaan"
                ws2[f'A{c + 14}'].value = "modus"
                ws2[f'A{c + 15}'].value = "steekproefvariatie"
                ws2[f'A{c + 16}'].value = "kurtosis"
                ws2[f'A{c + 17}'].value = "scheefheid"
                ws2[f'A{c + 18}'].value = "bereik"
                ws2[f'A{c + 19}'].value = "minimum"
                ws2[f'A{c + 20}'].value = "maximum"
                ws2[f'A{c + 21}'].value = "som"
                ws2[f'A{c + 22}'].value = "aantal stalen"
                ws2['P1'] = "creatinine outliers"
              
                merg = f'A{c + 8}:B{c + 8}'
                
                ws2.merge_cells(merg)
    
                


                for index, row in enumerate(ws2.iter_rows(values_only=True), start=1):
                    if index > 1 and index < gemiddeld:
            # Zorg ervoor dat row[14] een geldige waarde heeft
                        if row[14] is not None and isinstance(row[14], (int, float)):
                            if row[14] < 0.50:
                                ws2[f'P{index}'].value = "*"
                                if ws2[f'A{c + 2}'].value is None:
                                    ws2[f'A{c + 2}'].value = "*Gezien urine creatinine <0,50 dienen resultaten uitgedrukt per gram creatinine onder voorbehoud geïnterpreteerd te worden."
                            elif row[14] > 3.00:
                                ws2[f'P{index}'].value = "**"
                                if ws2[f'A{c + 3}'].value is None:
                                    ws2[f'A{c + 3}'].value = "**Gezien urine creatinine >3,00 dienen resultaten uitgedrukt per gram creatinine onder voorbehoud geïnterpreteerd te worden."
                
                ws2[f'A{c + 5}'].value = f'Detectielimiet: {value_M1}'
                ws2[f'A{c + 6}'].value = f'Resultaten kleiner als "<" werden als {minima} gerekend'
                for column in ws2.columns:
                        max_length = 0
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except TypeError:  # Als de celwaarde geen string is
                                pass
                        adjusted_width = (max_length + 2)
                        ws2.column_dimensions[column[0].column_letter].width = adjusted_width

                column_letter = 'A'
                column_letter_C = 'C'
                column_letter_I = 'I'
                new_width = 17.56
                new_width_C = 17.00
                new_width_I = 17.00
                ws2.column_dimensions[column_letter].width = new_width
                ws2.column_dimensions[column_letter_C].width = new_width_C
                ws2.column_dimensions[column_letter_I].width = new_width_I
                














        
        wb_schrijven.save("statistics.xlsx")
        with open("statistics.xlsx", "rb") as file:
            btn = st.download_button(
                label=f"Klik hier om {file_name}.xlsx te downloaden",
                data=file,
                file_name=f"{file_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

        
        st.success(f"Rijen met waarde '{zoekwaarde}' zijn gekopieerd naar statistics.xlsx.")



else: 
    st.error("Een of meer van de benodigde bestanden zijn niet geüpload. Controleer de bestanden en probeer het opnieuw.")
